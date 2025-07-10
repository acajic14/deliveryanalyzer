import sys
import os

if getattr(sys, 'frozen', False):
    sys.argv = [sys.argv[0], "run"]
    os.environ['STREAMLIT_RUNNING_VIA_PYINSTALLER'] = 'true'
    os.environ['STREAMLIT_SERVER_ENABLE_STATIC_SERVE'] = 'true'
    os.environ['STREAMLIT_SERVER_ENABLE_XSRF_PROTECTION'] = 'false'

import streamlit as st
import pandas as pd
from datetime import datetime
import re
from rapidfuzz import fuzz, process
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# --- Helper functions ---
def normalize_diacritics(text):
    diacritic_map = {'č':'c', 'š':'s', 'ž':'z', 'Č':'C', 'Š':'S', 'Ž':'Z'}
    return ''.join(diacritic_map.get(c, c) for c in text)

def clean_city_name(city):
    city = str(city).split('-')[0].split('–')[0].strip()
    return normalize_diacritics(city).lower()

def clean_street_name(address):
    irrelevant_words = {
        'slovenia', 'slovenija', 'slo', 'ljubljana', 'lj', 'avenija',
        'ulica', 'cesta', 'ul', 'street', 'road', 'd.o.o.', 'd.d.', 'eu', 'skl', 'vh', 'naselje', 'mesto'
    }
    address = normalize_diacritics(str(address))
    address = re.sub(r'[^\w\s]', ' ', address)
    address = re.sub(r'\s+', ' ', address).lower()
    address = re.sub(r'([a-z])(\d)', r'\1 \2', address)
    address = re.sub(r'(\d)([a-z])', r'\1 \2', address)
    parts = address.split()
    while parts and (re.match(r'^\d{4}$', parts[-1]) or parts[-1] in irrelevant_words):
        parts.pop()
    cleaned_parts = []
    for part in parts:
        if part in irrelevant_words or len(part) <= 2:
            continue
        cleaned_parts.append(part)
        if re.match(r'^\d+[a-z]?$', part):
            break
    return ' '.join(cleaned_parts).strip()

def normalize_consignee_name(name):
    if not isinstance(name, str): return ''
    name = name.lower().strip()
    name = re.sub(r'[.,]', '', name)
    suffixes = [
        'doo', 'd o o', 'd.o.o', 'd.o.o.', 'd d', 'd.d.', 'd.d', 'dno', 'd.n.o.', 'd.n.o', 'ddoo', 
        'ltd', 'limited', 'llc', 'gmbh', 'inc'
    ]
    for suffix in suffixes:
        name = re.sub(r'\s*' + re.escape(suffix) + r'$', '', name)
    return re.sub(r'\s+', ' ', name).strip()

def extract_house_number(address):
    match = re.search(r'\b(\d+[a-z]?)\b', str(address).lower())
    return match.group(1) if match else None

def house_number_to_float(hn):
    if not hn: return None
    hn = str(hn).lower()
    num_part = re.sub(r'[^0-9]', '', hn)
    letter_part = re.sub(r'[^a-z]', '', hn)
    if not num_part: return None
    base = float(num_part)
    if letter_part: base += (ord(letter_part) - ord('a') + 1) * 0.01
    return base

def load_street_city_routes(path):
    try:
        if not os.path.exists(path):
            st.warning(f"⚠️ Street-city routes file not found: {path}")
            return pd.DataFrame(columns=['ROUTE', 'STREET', 'CITY', 'CITY_CLEAN', 'STREET_CLEAN'])
        
        df = pd.read_excel(path)
        df.columns = ['ROUTE', 'STREET', 'CITY']
        df['CITY_CLEAN'] = df['CITY'].apply(clean_city_name)
        df['STREET_CLEAN'] = df['STREET'].apply(clean_street_name)
        return df
    except Exception as e:
        st.error(f"❌ Street-city routes error: {str(e)}")
        return pd.DataFrame(columns=['ROUTE', 'STREET', 'CITY', 'CITY_CLEAN', 'STREET_CLEAN'])

def load_fallback_routes(path):
    try:
        if not os.path.exists(path):
            st.warning(f"⚠️ Fallback routes file not found: {path}")
            return pd.DataFrame(columns=['ROUTE', 'ZIP'])
        
        df = pd.read_excel(path)
        df.columns = ['ROUTE', 'ZIP']
        df['ZIP'] = df['ZIP'].astype(str).str.zfill(4)
        return df
    except Exception as e:
        st.error(f"❌ Fallback routes error: {str(e)}")
        return pd.DataFrame(columns=['ROUTE', 'ZIP'])

def load_targets(path):
    try:
        if not os.path.exists(path):
            st.warning(f"⚠️ Targets file not found: {path}")
            return pd.DataFrame(columns=['ROUTE', 'Average PU stops', 'Target stops', 'SERVICE_PARTNER', 'LIST_OF_SP', 'SUM_TARGET_STOPS', 'AVG_ROUTES', 'SPR'])
        
        # Load WITHOUT headers to get raw data
        df = pd.read_excel(path, header=None)
        
        # Skip the header row (row 0) and start from row 1
        df_data = df.iloc[1:].reset_index(drop=True)
        
        # Create explicit column mapping by position
        if len(df.columns) >= 11:
            df_mapped = pd.DataFrame()
            df_mapped['ROUTE'] = df_data.iloc[:, 0].astype(str)                    # Column A
            df_mapped['Average PU stops'] = pd.to_numeric(df_data.iloc[:, 1], errors='coerce').fillna(0)        # Column B  
            df_mapped['Target stops'] = pd.to_numeric(df_data.iloc[:, 2], errors='coerce').fillna(0)            # Column C
            df_mapped['SERVICE_PARTNER'] = df_data.iloc[:, 4].astype(str)         # Column E - Individual route SP
            df_mapped['LIST_OF_SP'] = df_data.iloc[:, 7].astype(str)              # Column H - Unique SP list
            df_mapped['SUM_TARGET_STOPS'] = pd.to_numeric(df_data.iloc[:, 8], errors='coerce').fillna(0)        # Column I
            df_mapped['AVG_ROUTES'] = pd.to_numeric(df_data.iloc[:, 9], errors='coerce').fillna(0)              # Column J
            df_mapped['SPR'] = pd.to_numeric(df_data.iloc[:, 10], errors='coerce').fillna(0)                    # Column K
            
            # Remove rows where ROUTE is NaN or empty
            df_mapped = df_mapped[df_mapped['ROUTE'].notna()]
            df_mapped = df_mapped[df_mapped['ROUTE'] != 'nan']
            df_mapped = df_mapped[df_mapped['ROUTE'] != '']
            
            # Check if SERVICE_PARTNER column has valid data
            non_null_sp = df_mapped[df_mapped['SERVICE_PARTNER'].notna() & (df_mapped['SERVICE_PARTNER'] != 'nan') & (df_mapped['SERVICE_PARTNER'] != '')]
            
            if len(non_null_sp) == 0:
                st.error("❌ Column E (SERVICE_PARTNER) appears to be empty after processing!")
            else:
                st.success(f"✅ Found {len(non_null_sp)} routes with service partner assignments")
            
            return df_mapped
        else:
            st.error(f"❌ Targets file needs at least 11 columns, found {len(df.columns)}")
            return pd.DataFrame(columns=['ROUTE', 'Average PU stops', 'Target stops', 'SERVICE_PARTNER', 'LIST_OF_SP', 'SUM_TARGET_STOPS', 'AVG_ROUTES', 'SPR'])
        
    except Exception as e:
        st.error(f"❌ Couldn't load targets.xlsx: {str(e)}")
        return pd.DataFrame(columns=['ROUTE', 'Average PU stops', 'Target stops', 'SERVICE_PARTNER', 'LIST_OF_SP', 'SUM_TARGET_STOPS', 'AVG_ROUTES', 'SPR'])

def calculate_service_partner_spr(route_summary, targets_df):
    """Calculate predicted SPR per service partner with route adjustment recommendations"""
    try:
        if targets_df.empty or 'SERVICE_PARTNER' not in targets_df.columns:
            return pd.DataFrame()
        
        # Direct merge on exact route names (CE1A = CE1A)
        merged = pd.merge(
            route_summary[['ROUTE', 'Predicted Stops']], 
            targets_df[['ROUTE', 'SERVICE_PARTNER']], 
            on='ROUTE',
            how='left'
        )
        
        # Remove rows where SERVICE_PARTNER is NaN or empty
        merged = merged.dropna(subset=['SERVICE_PARTNER'])
        merged = merged[merged['SERVICE_PARTNER'] != '']
        merged = merged[merged['SERVICE_PARTNER'] != 'nan']
        
        if merged.empty:
            return pd.DataFrame()
        
        # Group by service partner to sum predicted stops
        predicted_spr_sp = merged.groupby('SERVICE_PARTNER').agg({
            'Predicted Stops': 'sum'
        }).reset_index()
        
        # Get unique service partner info from targets
        service_partners = targets_df[['LIST_OF_SP', 'SUM_TARGET_STOPS', 'AVG_ROUTES', 'SPR']].copy()
        service_partners = service_partners.dropna(subset=['LIST_OF_SP'])
        service_partners = service_partners[service_partners['LIST_OF_SP'] != '']
        service_partners = service_partners[service_partners['LIST_OF_SP'] != 'nan']
        service_partners = service_partners.drop_duplicates(subset=['LIST_OF_SP'])
        
        # Merge predicted SPR with service partner info
        spr_summary = pd.merge(
            service_partners, 
            predicted_spr_sp, 
            left_on='LIST_OF_SP', 
            right_on='SERVICE_PARTNER', 
            how='left'
        )
        
        if spr_summary.empty:
            return pd.DataFrame()
        
        # Calculate percent of predicted stops vs target
        spr_summary['Percent_of_Target'] = (spr_summary['Predicted Stops'] / spr_summary['SUM_TARGET_STOPS']) * 100
        
        # Calculate predicted SPR per service partner
        spr_summary['Predicted_SPR'] = spr_summary['Predicted Stops'] / spr_summary['AVG_ROUTES']

        # Calculate routes to add or remove
        # Formula: (Predicted Stops - Target Stops) / SPR Target
        # Positive value means routes to add, negative means routes to remove
        spr_summary['Routes_to_Adjust'] = (spr_summary['Predicted Stops'] - spr_summary['SUM_TARGET_STOPS']) / spr_summary['SPR']

        # Round values for display
        spr_summary['Predicted Stops'] = spr_summary['Predicted Stops'].fillna(0).round(1)
        spr_summary['Percent_of_Target'] = spr_summary['Percent_of_Target'].fillna(0).round(1)
        spr_summary['Predicted_SPR'] = spr_summary['Predicted_SPR'].fillna(0).round(1)
        spr_summary['SUM_TARGET_STOPS'] = spr_summary['SUM_TARGET_STOPS'].fillna(0).astype(int)
        spr_summary['AVG_ROUTES'] = spr_summary['AVG_ROUTES'].fillna(0).astype(int)
        spr_summary['SPR'] = spr_summary['SPR'].fillna(0).round(1)
        spr_summary['Routes_to_Adjust'] = spr_summary['Routes_to_Adjust'].fillna(0).round(1)
        
        # Clean up columns
        spr_summary = spr_summary[['LIST_OF_SP', 'SUM_TARGET_STOPS', 'AVG_ROUTES', 'SPR', 'Predicted Stops', 'Predicted_SPR', 'Percent_of_Target', 'Routes_to_Adjust']]
        spr_summary.columns = ['Service Partner', 'Target Stops', 'Avg Routes', 'SPR Target', 'Predicted Stops', 'Predicted SPR', 'Percent of Target (%)', 'Routes to Add/Remove']
        
        return spr_summary
        
    except Exception as e:
        st.error(f"Error calculating service partner SPR: {str(e)}")
        return pd.DataFrame()

def add_service_partner_spr_summary(workbook, sheet, spr_summary):
    """Add service partner SPR summary AFTER DHL headers"""
    if spr_summary.empty:
        return 0
    
    # Find where to insert (after DHL headers which are 4 rows)
    insert_row = 5  # After DHL headers
    
    # Insert rows for the summary
    rows_needed = len(spr_summary) + 3  # title + header + data
    sheet.insert_rows(insert_row, rows_needed)
    
    # DHL colors
    dhl_yellow = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
    dhl_red = PatternFill(start_color="D40511", end_color="D40511", fill_type="solid")
    
    # Title
    title_row = insert_row
    sheet.merge_cells(f'A{title_row}:H{title_row}')
    sheet[f'A{title_row}'] = 'DHL Service Partner SPR Summary'
    sheet[f'A{title_row}'].font = Font(bold=True, size=14, color="FFFFFF")
    sheet[f'A{title_row}'].fill = dhl_red
    sheet[f'A{title_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    header_row = title_row + 1
    headers = spr_summary.columns.tolist()
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=header_row, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = dhl_yellow
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    data_start_row = header_row + 1
    for idx, row in spr_summary.iterrows():
        for col_num, value in enumerate(row, 1):
            cell = sheet.cell(row=data_start_row + idx, column=col_num, value=value)
            # Highlight percentage column based on performance
            if col_num == len(headers) - 1:  # Percent of Target column
                if value >= 100:
                    cell.font = Font(color="008000", bold=True)  # Green for over target
                elif value >= 90:
                    cell.font = Font(color="FFA500", bold=True)  # Orange for close to target
                else:
                    cell.font = Font(color="FF0000", bold=True)  # Red for under target
            # Highlight route adjustment column
            elif col_num == len(headers):  # Routes to Add/Remove column
                if value < -0.5:
                    cell.font = Font(color="FF0000", bold=True)  # Red for remove routes
                elif value > 0.5:
                    cell.font = Font(color="008000", bold=True)  # Green for add routes
                else:
                    cell.font = Font(color="FFA500", bold=True)  # Orange for optimal
    
    return rows_needed

def parse_pieces(value):
    try:
        return int(re.findall(r'\d+', str(value).split('/')[0])[0])
    except:
        return 1

def clean_nan_from_address(addr):
    return re.sub(r'\s*nan\s*$', '', str(addr)).strip() if isinstance(addr, str) else addr

def load_email_mapping(path):
    try:
        if not os.path.exists(path):
            st.warning(f"⚠️ Email mapping file not found: {path}")
            return pd.DataFrame(columns=['Report_Type', 'Email', 'Contact_Name'])
        
        return pd.read_excel(path)
    except Exception as e:
        st.warning(f"⚠️ Couldn't load email mapping: {str(e)}")
        return pd.DataFrame(columns=['Report_Type', 'Email', 'Contact_Name'])

def add_dhl_branding_to_excel(workbook, worksheet, title="DHL Route Analysis"):
    """Add enhanced DHL branding to Excel worksheets"""
    # DHL Colors
    dhl_yellow = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
    dhl_red = PatternFill(start_color="D40511", end_color="D40511", fill_type="solid")
    
    # Add DHL header rows
    worksheet.insert_rows(1, 4)  # Insert 4 rows
    
    # Merge cells for header
    worksheet.merge_cells('A1:J1')
    worksheet.merge_cells('A2:J2')
    worksheet.merge_cells('A3:J3')
    
    # Set header content with enhanced branding
    worksheet['A1'] = "DHL EXPRESS"
    worksheet['A2'] = "Excellence. Simply delivered."
    worksheet['A3'] = title
    worksheet['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | DHL Route Optimization System"
    
    # Style header
    worksheet['A1'].fill = dhl_red
    worksheet['A1'].font = Font(color="FFFFFF", bold=True, size=18)
    worksheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    worksheet['A2'].fill = dhl_yellow
    worksheet['A2'].font = Font(color="000000", bold=True, size=12, italic=True)
    worksheet['A2'].alignment = Alignment(horizontal="center", vertical="center")
    
    worksheet['A3'].fill = dhl_yellow
    worksheet['A3'].font = Font(color="D40511", bold=True, size=14)
    worksheet['A3'].alignment = Alignment(horizontal="center", vertical="center")
    
    worksheet['A4'].font = Font(color="666666", size=9)
    worksheet['A4'].alignment = Alignment(horizontal="center")
    
    # Adjust row heights
    worksheet.row_dimensions[1].height = 30
    worksheet.row_dimensions[2].height = 18
    worksheet.row_dimensions[3].height = 22
    worksheet.row_dimensions[4].height = 15

def send_email_with_attachment(smtp_server, smtp_port, sender_email, sender_password, 
                              recipient_email, contact_name, subject, body, attachment_path):
    try:
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = subject
        
        # DHL branded email body
        dhl_body = f"""
DHL EXPRESS - Excellence. Simply delivered.

{body}

---
This report was generated by the DHL Route Analyzer System.
For support, please contact your DHL operations team.

© 2025 DHL International GmbH
Excellence. Simply delivered.
"""
        
        msg.attach(MIMEText(dhl_body, 'plain'))
        
        if os.path.exists(attachment_path):
            with open(attachment_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {os.path.basename(attachment_path)}'
            )
            msg.attach(part)
        
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()
        
        return True, f"✅ Email sent successfully to {contact_name} ({recipient_email})"
    
    except Exception as e:
        return False, f"❌ Failed to send email to {recipient_email}: {str(e)}"

def send_route_reports(route_summary, specialized_reports, email_mapping, output_path, timestamp,
                      smtp_server, smtp_port, sender_email, sender_password):
    results = []
    email_routes = {
        'MBX': ['MB1', 'MB2'],
        'KRA': ['KR1', 'KR2'], 
        'LJU': ['LJ1', 'LJ2'],
        'NMO': ['NM1', 'NM2'],
        'CEJ': ['CE1', 'CE2'],
        'NGR': ['NG1', 'NG2'],
        'NGX': ['NGX'],
        'KOP': ['KP1']
    }
    for report_type, route_prefixes in email_routes.items():
        recipients = email_mapping[email_mapping['Report_Type'] == report_type]
        if not recipients.empty and report_type in specialized_reports:
            attachment_path = specialized_reports[report_type]
            route_data = route_summary[route_summary['ROUTE'].str.startswith(tuple(route_prefixes), na=False)]
            total_shipments = route_data['total_shipments'].sum() if not route_data.empty else 0
            total_weight = route_data['total_weight'].sum() if not route_data.empty else 0
            subject = f"DHL Express Daily Route Report - {report_type} ({datetime.now().strftime('%Y-%m-%d')})"
            body = f"""Dear Team,

Please find attached the daily route report for {report_type} routes ({', '.join(route_prefixes)}).

Summary for {datetime.now().strftime('%Y-%m-%d')}:
- Total Shipments: {total_shipments}
- Total Weight: {total_weight:.1f} kg
- Routes Covered: {', '.join(route_prefixes)}

The attached Excel file contains detailed shipment information including:
- Route assignments
- Consignee details
- Addresses and ZIP codes
- AWB numbers

Please review and coordinate accordingly.

Best regards,
DHL Route Analyzer System
"""
            for _, recipient in recipients.iterrows():
                success, message = send_email_with_attachment(
                    smtp_server, smtp_port, sender_email, sender_password,
                    recipient['Email'], recipient['Contact_Name'],
                    subject, body, attachment_path
                )
                results.append((report_type, recipient['Contact_Name'], success, message))
    return results
def apply_column_mapping(df):
    column_map = {
        'HWB': 'HWB',
        '# Pcs\\Tot Pcs': 'PIECES',
        'Wt': 'WEIGHT',
        'Vol Wt': 'VOLUMETRIC_WEIGHT',
        'Cnee Nm': 'CONSIGNEE_NAME',
        'Cnee Addr Ln 1': 'CONSIGNEE_STREET1',
        'Cnee Addr Ln 2': 'CONSIGNEE_STREET2',
        'Cnee Zip': 'CONSIGNEE_ZIP',
        'Cnee City': 'CONSIGNEE_CITY',
        'Cnee Str #': 'HOUSE_NUMBER',
        'PCC': 'PCC'
    }
    new_df = pd.DataFrame()
    for orig, new in column_map.items():
        if orig in df.columns:
            new_df[new] = df[orig].astype(str)
        else:
            new_df[new] = pd.Series('', index=df.index, dtype='str')
    new_df['PIECES'] = new_df['PIECES'].apply(parse_pieces)
    new_df['PCC'] = new_df['PCC'].astype(str).str.strip().str.upper() if 'PCC' in new_df.columns else None
    street1 = new_df['CONSIGNEE_STREET1'].fillna('')
    street2 = new_df['CONSIGNEE_STREET2'].fillna('')
    new_df['CONSIGNEE_STREET'] = street1.str.strip() + ' ' + street2.str.strip()
    new_df['CONSIGNEE_STREET'] = new_df['CONSIGNEE_STREET'].str.replace('nan', '').str.strip()
    new_df['HOUSE_NUMBER'] = new_df['CONSIGNEE_STREET'].apply(extract_house_number)
    new_df['HOUSE_NUMBER_FLOAT'] = new_df['HOUSE_NUMBER'].apply(house_number_to_float)
    new_df['STREET_NAME'] = new_df['CONSIGNEE_STREET'].apply(clean_street_name)
    if 'CONSIGNEE_ZIP' in new_df.columns:
        new_df['CONSIGNEE_ZIP'] = new_df['CONSIGNEE_ZIP'].astype(str).str.extract(r'(\d{4})')[0].str.zfill(4)
    for col in ['WEIGHT', 'VOLUMETRIC_WEIGHT']:
        if col in new_df.columns: 
            new_df[col] = pd.to_numeric(new_df[col], errors='coerce').fillna(0)
    new_df['PIECES'] = pd.to_numeric(new_df['PIECES'], errors='coerce').fillna(1)
    new_df['MATCHED_ROUTE'] = None
    new_df['MATCH_SCORE'] = 0.0
    new_df['MATCH_METHOD'] = None
    new_df['CONSIGNEE_ADDRESS'] = new_df['CONSIGNEE_STREET'].apply(clean_nan_from_address)
    new_df['CONSIGNEE_NAME_NORM'] = new_df['CONSIGNEE_NAME'].apply(normalize_consignee_name)
    return new_df

def process_multiple_manifests(uploaded_files):
    all_dataframes = []
    for i, file in enumerate(uploaded_files):
        st.write(f"Processing file {i+1}/{len(uploaded_files)}: {file.name}")
        try:
            if file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file, header=0)
            else:
                df = pd.read_csv(file, delimiter=',', quotechar='"', engine='python', header=0, dtype=str)
            df = apply_column_mapping(df)
            if not df.empty:
                df = df[df['HWB'] != 'HWB']
                df = df[df['HWB'].notna()]
                df = df[df['HWB'] != '']
                df['SOURCE_FILE'] = file.name
                df['FILE_ORDER'] = i + 1
                all_dataframes.append(df)
                st.success(f"✅ {file.name}: {len(df)} rows processed")
            else:
                st.warning(f"⚠️ {file.name}: No data found")
        except Exception as e:
            st.error(f"❌ {file.name}: Error - {str(e)}")
            continue
    if not all_dataframes:
        st.error("❌ No valid data found in any uploaded files")
        return pd.DataFrame()
    merged_df = pd.concat(all_dataframes, ignore_index=True)
    merged_df = merged_df[merged_df['HWB'] != 'HWB']
    merged_df = merged_df[merged_df['CONSIGNEE_NAME'] != 'Cnee Nm']
    merged_df = merged_df.reset_index(drop=True)
    
    # DHL branded success message
    st.markdown("""
    <div style="background: linear-gradient(90deg, #FFCC00 0%, #D40511 100%); 
                padding: 1rem; border-radius: 10px; text-align: center; margin: 1rem 0;">
        <h3 style="color: white; margin: 0;">🎉 DHL File Merge Complete!</h3>
        <p style="color: white; margin: 0;">Excellence. Simply delivered.</p>
    </div>
    """, unsafe_allow_html=True)
    
    st.write(f"- **Total rows:** {len(merged_df)}")
    st.write(f"- **Files merged:** {len(all_dataframes)}")
    st.write(f"- **Unique HWBs:** {merged_df['HWB'].nunique()}")
    with st.expander("📊 File Breakdown"):
        file_summary = merged_df.groupby('SOURCE_FILE').agg({
            'HWB': 'count',
            'WEIGHT': 'sum',
            'PIECES': 'sum'
        }).round(1)
        file_summary.columns = ['Rows', 'Total Weight', 'Total Pieces']
        st.dataframe(file_summary)
    return merged_df

def process_manifest(file):
    try:
        if file.name.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file)
        else:
            df = pd.read_csv(file, delimiter=',', quotechar='"', engine='python', dtype=str)
    except Exception as e:
        st.error(f"❌ Error reading file: {str(e)}")
        return pd.DataFrame()
    return apply_column_mapping(df)

def match_address_to_route(manifest_df, street_city_routes, fallback_routes):
    # Ensure DataFrames are initialized
    if street_city_routes is None:
        street_city_routes = pd.DataFrame(columns=['ROUTE', 'STREET', 'CITY', 'CITY_CLEAN', 'STREET_CLEAN'])
    if fallback_routes is None:
        fallback_routes = pd.DataFrame(columns=['ROUTE', 'ZIP'])
    
    manifest_df['MATCHED_ROUTE'] = None
    manifest_df['MATCH_METHOD'] = None
    manifest_df['MATCH_SCORE'] = 0.0
    special_zips = {'2000', '3000', '4000', '5000', '6000', '8000'}

    for idx, row in manifest_df.iterrows():
        consignee_name = str(row['CONSIGNEE_NAME']).lower()
        consignee_street = str(row['CONSIGNEE_STREET']).lower()
        if ('elrad' in consignee_name and 'electronics' in consignee_name and 
            ('ljutomerska' in consignee_street and '47' in consignee_street)):
            manifest_df.at[idx, 'MATCHED_ROUTE'] = 'MB1B'
            manifest_df.at[idx, 'MATCH_METHOD'] = 'Direct Rule - Elrad'
            manifest_df.at[idx, 'MATCH_SCORE'] = 100.0
            continue
        
        zip_code = row['CONSIGNEE_ZIP']
        street_name = clean_street_name(row['CONSIGNEE_STREET'])
        city_name = clean_city_name(row['CONSIGNEE_CITY'])
        matched = False

        # Check if we have valid route data
        has_street_city = not street_city_routes.empty
        has_fallback = not fallback_routes.empty

        if zip_code in special_zips:
            if has_street_city:
                city_matches = street_city_routes[street_city_routes['CITY_CLEAN'] == city_name]
                if not city_matches.empty:
                    matches = process.extract(street_name, city_matches['STREET_CLEAN'], scorer=fuzz.token_set_ratio, score_cutoff=70, limit=3)
                    for matched_street, score, _ in matches:
                        best_match = city_matches[city_matches['STREET_CLEAN'] == matched_street].iloc[0]
                        manifest_df.at[idx, 'MATCHED_ROUTE'] = best_match['ROUTE']
                        manifest_df.at[idx, 'MATCH_METHOD'] = 'Street-City'
                        manifest_df.at[idx, 'MATCH_SCORE'] = float(score)
                        matched = True
                        break
            if not matched and has_fallback and zip_code in fallback_routes['ZIP'].values:
                manifest_df.at[idx, 'MATCHED_ROUTE'] = fallback_routes.loc[fallback_routes['ZIP'] == zip_code, 'ROUTE'].values[0]
                manifest_df.at[idx, 'MATCH_METHOD'] = 'ZIP'
                manifest_df.at[idx, 'MATCH_SCORE'] = 100.0
        else:
            if has_fallback and zip_code in fallback_routes['ZIP'].values:
                manifest_df.at[idx, 'MATCHED_ROUTE'] = fallback_routes.loc[fallback_routes['ZIP'] == zip_code, 'ROUTE'].values[0]
                manifest_df.at[idx, 'MATCH_METHOD'] = 'ZIP'
                manifest_df.at[idx, 'MATCH_SCORE'] = 100.0
            elif has_street_city:
                city_matches = street_city_routes[street_city_routes['CITY_CLEAN'] == city_name]
                if not city_matches.empty:
                    matches = process.extract(street_name, city_matches['STREET_CLEAN'], scorer=fuzz.token_set_ratio, score_cutoff=70, limit=3)
                    for matched_street, score, _ in matches:
                        best_match = city_matches[city_matches['STREET_CLEAN'] == matched_street].iloc[0]
                        manifest_df.at[idx, 'MATCHED_ROUTE'] = best_match['ROUTE']
                        manifest_df.at[idx, 'MATCH_METHOD'] = 'Street-City'
                        manifest_df.at[idx, 'MATCH_SCORE'] = float(score)
                        matched = True
                        break
    return manifest_df

def add_target_conditional_formatting(sheet, col_letter, start_row, end_row):
    red_font = Font(color='FF0000')
    yellow_font = Font(color='FFA500')
    green_font = Font(color='008000')
    sheet.conditional_formatting.add(f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='greaterThan', formula=['5'], font=red_font))
    sheet.conditional_formatting.add(f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='lessThan', formula=['-5'], font=red_font))
    sheet.conditional_formatting.add(f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='between', formula=['-5', '0'], font=yellow_font))
    sheet.conditional_formatting.add(f'{col_letter}{start_row}:{col_letter}{end_row}',
        CellIsRule(operator='between', formula=['0', '5'], font=green_font))

def auto_adjust_column_width(worksheet):
    from openpyxl.cell.cell import MergedCell
    for column in worksheet.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if isinstance(cell, MergedCell):
                continue
            if column_letter is None:
                column_letter = cell.column_letter
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

def create_specialized_report(manifest_df, route_prefixes, report_name, output_path, timestamp):
    filtered_data = manifest_df[
        manifest_df['MATCHED_ROUTE'].str.startswith(tuple(route_prefixes), na=False)
    ].copy()
    report_path = f"{output_path}/{report_name}_details_{timestamp}.xlsx"
    try:
        if not filtered_data.empty:
            report_data = filtered_data[['MATCHED_ROUTE', 'HWB', 'CONSIGNEE_NAME', 'CONSIGNEE_ADDRESS', 'CONSIGNEE_ZIP']].copy()
            report_data = report_data.sort_values(by=['MATCHED_ROUTE', 'CONSIGNEE_ZIP'], ascending=[True, True])
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                report_data.to_excel(writer, index=False, startrow=5)
                workbook = writer.book
                sheet = writer.sheets['Sheet1']
                add_dhl_branding_to_excel(workbook, sheet, f"DHL {report_name} Route Details")
                auto_adjust_column_width(sheet)
        else:
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                pd.DataFrame(columns=['MATCHED_ROUTE', 'HWB', 'CONSIGNEE_NAME', 'CONSIGNEE_ADDRESS', 'CONSIGNEE_ZIP']).to_excel(writer, index=False, startrow=5)
                workbook = writer.book
                sheet = writer.sheets['Sheet1']
                add_dhl_branding_to_excel(workbook, sheet, f"DHL {report_name} Route Details")
        st.write(f"✅ Created specialized report: {report_name}")
    except Exception as e:
        st.error(f"❌ Failed to create {report_name} report: {str(e)}")
    return report_path

def identify_multi_shipment_customers(manifest_df):
    customer_shipments = manifest_df.groupby('CONSIGNEE_NAME_NORM').agg(
        total_shipments=('HWB', 'nunique'),
        total_pieces=('PIECES', 'sum'),
        zip_code=('CONSIGNEE_ZIP', 'first')
    ).reset_index()
    multi_shipment_customers = customer_shipments[customer_shipments['total_shipments'] > 1]
    return multi_shipment_customers.sort_values(
        by=['zip_code', 'total_shipments'],
        ascending=[True, False]
    )

def generate_reports(
    manifest_df, output_path, weight_thr=70, vol_weight_thr=150, pieces_thr=6,
    vehicle_weight_thr=70, vehicle_vol_thr=150, vehicle_pieces_thr=12,
    vehicle_kg_per_piece_thr=10, vehicle_van_max_pieces=20, multi_shipment_thr=5
):
    # Initialize specialized_reports early to avoid UnboundLocalError
    specialized_reports = {}
    
    st.write("🔄 Starting report generation...")
    
    # Ensure required columns exist
    required_cols = ['HWB', 'MATCHED_ROUTE', 'CONSIGNEE_NAME_NORM', 'CONSIGNEE_NAME', 
                    'CONSIGNEE_ZIP', 'CONSIGNEE_ADDRESS', 'WEIGHT', 'VOLUMETRIC_WEIGHT', 'PIECES', 'CONSIGNEE_CITY']
    
    for col in required_cols:
        if col not in manifest_df.columns:
            st.error(f"❌ Missing required column: {col}")
            manifest_df[col] = ""

    hwb_aggregated = manifest_df.groupby(['HWB', 'MATCHED_ROUTE']).agg({
        'CONSIGNEE_NAME_NORM': 'first',
        'CONSIGNEE_NAME': 'first',
        'CONSIGNEE_ZIP': 'first',
        'CONSIGNEE_ADDRESS': 'first',
        'WEIGHT': 'sum',
        'VOLUMETRIC_WEIGHT': 'sum',
        'PIECES': 'first'
    }).reset_index()

    route_summary = hwb_aggregated.groupby('MATCHED_ROUTE').agg(
        total_shipments=('HWB', 'count'),
        unique_consignees=('CONSIGNEE_NAME_NORM', 'nunique'),
        total_weight=('WEIGHT', 'sum'),
        total_pieces=('PIECES', 'sum')
    ).reset_index().rename(columns={'MATCHED_ROUTE': 'ROUTE'})

    targets_df = load_targets('input/targets.xlsx')
    if not targets_df.empty:
        route_summary = pd.merge(route_summary, targets_df, on='ROUTE', how='left')
    else:
        route_summary['Average PU stops'] = 0
        route_summary['Target stops'] = 0

    route_summary['total_weight'] = route_summary['total_weight'].round(1)
    route_summary['Predicted Stops'] = route_summary['unique_consignees'] + route_summary['Average PU stops']
    route_summary['Predicted - Target'] = route_summary['Predicted Stops'] - route_summary['Target stops']
    for col in ['Average PU stops', 'Predicted Stops', 'Predicted - Target']:
        route_summary[col] = route_summary[col].round(0).astype('Int64')
    route_summary.insert(5, '', '')
    route_summary = route_summary[['ROUTE', 'total_shipments', 'unique_consignees', 'total_weight', 'total_pieces', '',
                                  'Average PU stops', 'Predicted Stops', 'Target stops', 'Predicted - Target']]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    os.makedirs(output_path, exist_ok=True)
    
    summary_path = f"{output_path}/DHL_route_summary_{timestamp}.xlsx"
    special_cases_path = f"{output_path}/DHL_special_cases_{timestamp}.xlsx"
    matching_details_path = f"{output_path}/DHL_matching_details_{timestamp}.xlsx"
    wth_mpcs_path = f"{output_path}/DHL_WTH_MPCS_Report_{timestamp}.xlsx"
    priority_path = f"{output_path}/DHL_Priority_Shipments_{timestamp}.xlsx"
    multi_shipments_path = f"{output_path}/DHL_multi_shipments_{timestamp}.xlsx"

    # 1. ROUTE SUMMARY REPORT
    try:
        st.write("📊 Creating route summary report...")
        with pd.ExcelWriter(summary_path, engine='openpyxl') as writer:
            route_summary.to_excel(writer, sheet_name='Summary', index=False, startrow=5)
            workbook = writer.book
            sheet = writer.sheets['Summary']
            
            add_dhl_branding_to_excel(workbook, sheet, "Daily Route Summary Report")
            
            spr_summary = calculate_service_partner_spr(route_summary, targets_df)
            
            rows_added = 0
            if not spr_summary.empty:
                rows_added = add_service_partner_spr_summary(workbook, sheet, spr_summary)
            
            current_row = sheet.max_row + 2
            
            avg_predicted_stops = route_summary['Predicted Stops'].mean() if not route_summary.empty else 0
            unmatched_count = len(manifest_df[manifest_df['MATCHED_ROUTE'].isna() | (manifest_df['MATCHED_ROUTE'] == '')])
            
            sheet.cell(row=current_row, column=1, value="Average Predicted Stops")
            sheet.cell(row=current_row, column=2, value=round(avg_predicted_stops, 1))
            current_row += 1
            sheet.cell(row=current_row, column=1, value="Unmatched Route Shipments")
            sheet.cell(row=current_row, column=2, value=unmatched_count)
            current_row += 2
            
            sheet.cell(row=current_row, column=1, value="PCC Statistics:")
            current_row += 1
            sheet.cell(row=current_row, column=1, value="Product")
            sheet.cell(row=current_row, column=2, value="Shipments")
            sheet.cell(row=current_row, column=3, value="Pieces")
            sheet.cell(row=current_row, column=4, value="Pieces/Shipment")
            current_row += 1
            
            pcc_categories = [('WPX','WPX'), ('TDY','TDY'), ('ESI','ESI'), ('ECX','ECX'), ('ESU','ESU'), ('ALL','All volume')]
            for code, label in pcc_categories:
                if 'PCC' in manifest_df.columns:
                    filtered = manifest_df[manifest_df['PCC'] == code] if code != 'ALL' else manifest_df
                    try:
                        shipments = int(filtered['HWB'].nunique())
                        pieces = int(filtered['PIECES'].sum())
                        ratio = round(pieces/shipments, 2) if shipments > 0 else 0.0
                    except (TypeError, ZeroDivisionError, ValueError):
                        shipments, pieces, ratio = 0, 0, 0.0
                else:
                    shipments, pieces, ratio = 0, 0, 0.0
                sheet.cell(row=current_row, column=1, value=label)
                sheet.cell(row=current_row, column=2, value=shipments)
                sheet.cell(row=current_row, column=3, value=pieces)
                sheet.cell(row=current_row, column=4, value=ratio)
                current_row += 1

            zip_stats = manifest_df.groupby('CONSIGNEE_ZIP').agg(
                total_shipments=('HWB', 'nunique'),
                unique_consignees=('CONSIGNEE_NAME_NORM', 'nunique')
            ).reset_index().sort_values('CONSIGNEE_ZIP')
            
            current_row += 2
            sheet.cell(row=current_row, column=1, value="ZIP Code Statistics:")
            current_row += 1
            sheet.cell(row=current_row, column=1, value="ZIP Code")
            sheet.cell(row=current_row, column=2, value="Shipments")
            sheet.cell(row=current_row, column=3, value="Unique Consignees")
            current_row += 1
            for _, row in zip_stats.iterrows():
                sheet.cell(row=current_row, column=1, value=row['CONSIGNEE_ZIP'])
                sheet.cell(row=current_row, column=2, value=row['total_shipments'])
                sheet.cell(row=current_row, column=3, value=row['unique_consignees'])
                current_row += 1

            if not route_summary.empty:
                format_start_row = 6 + rows_added
                format_end_row = format_start_row + len(route_summary) - 1
                add_target_conditional_formatting(sheet, 'J', format_start_row, format_end_row)

            route_prefixes = ['KR', 'LJ', 'KP', 'NG', 'NM', 'CE', 'MB']
            for prefix in route_prefixes:
                prefix_data = manifest_df[
                    manifest_df['MATCHED_ROUTE'].str.startswith(prefix, na=False)
                ].copy()
                
                if not prefix_data.empty:
                    sheet_data = prefix_data[[
                        'MATCHED_ROUTE', 'CONSIGNEE_NAME', 'CONSIGNEE_ADDRESS', 
                        'CONSIGNEE_CITY', 'CONSIGNEE_ZIP', 'HWB', 'PIECES'
                    ]].copy()
                    
                    sheet_data.columns = [
                        'MATCHED ROUTE', 'CONSIGNEE', 'CONSIGNEE ADDRESS', 
                        'CITY', 'ZIP', 'AWB', 'PIECES'
                    ]
                    
                    sheet_data = sheet_data.sort_values(['MATCHED ROUTE', 'ZIP'])
                    
                    try:
                        sheet_data.to_excel(writer, sheet_name=prefix, index=False, startrow=5)
                        prefix_sheet = writer.sheets[prefix]
                        add_dhl_branding_to_excel(workbook, prefix_sheet, f"DHL {prefix} Route Details")
                        auto_adjust_column_width(prefix_sheet)
                    except Exception as e:
                        st.warning(f"Could not create sheet {prefix}: {str(e)}")
                else:
                    empty_df = pd.DataFrame(columns=[
                        'MATCHED ROUTE', 'CONSIGNEE', 'CONSIGNEE ADDRESS', 
                        'CITY', 'ZIP', 'AWB', 'PIECES'
                    ])
                    empty_df.to_excel(writer, sheet_name=prefix, index=False, startrow=5)
                    prefix_sheet = writer.sheets[prefix]
                    add_dhl_branding_to_excel(workbook, prefix_sheet, f"DHL {prefix} Route Details")

            auto_adjust_column_width(sheet)
        st.success("✅ Route summary report created")
    except Exception as e:
        st.error(f"❌ Failed to create route summary: {str(e)}")

    # 2. SPECIALIZED REPORTS
    st.write("🚛 Creating specialized route reports...")
    specialized_reports['MBX'] = create_specialized_report(manifest_df, ['MB1', 'MB2'], 'MBX', output_path, timestamp)
    specialized_reports['KRA'] = create_specialized_report(manifest_df, ['KR1', 'KR2'], 'KRA', output_path, timestamp)
    specialized_reports['LJU'] = create_specialized_report(manifest_df, ['LJ1', 'LJ2'], 'LJU', output_path, timestamp)
    specialized_reports['NMO'] = create_specialized_report(manifest_df, ['NM1', 'NM2'], 'NMO', output_path, timestamp)
    specialized_reports['CEJ'] = create_specialized_report(manifest_df, ['CE1', 'CE2'], 'CEJ', output_path, timestamp)
    specialized_reports['NGR'] = create_specialized_report(manifest_df, ['NG1', 'NG2'], 'NGR', output_path, timestamp)
    specialized_reports['NGX'] = create_specialized_report(manifest_df, ['NGX'], 'NGX', output_path, timestamp)
    specialized_reports['KOP'] = create_specialized_report(manifest_df, ['KP1'], 'KOP', output_path, timestamp)

    # 3. SPECIAL CASES REPORT
    try:
        st.write("⚠️ Creating special cases report...")
        threshold_special_cases = manifest_df[
            (manifest_df['WEIGHT'] > weight_thr) |
            (manifest_df['VOLUMETRIC_WEIGHT'] > vol_weight_thr) |
            (manifest_df['PIECES'] > pieces_thr)
        ].copy()
        
        customer_shipments = manifest_df.groupby('CONSIGNEE_NAME_NORM').agg(
            total_shipments=('HWB', 'nunique'),
            total_pieces=('PIECES', 'sum'),
            zip_code=('CONSIGNEE_ZIP', 'first'),
            consignee_name=('CONSIGNEE_NAME', 'first'),
            matched_route=('MATCHED_ROUTE', 'first')
        ).reset_index()
        
        multi_shipment_special = customer_shipments[customer_shipments['total_shipments'] >= multi_shipment_thr].copy()
        
        # Prepare special cases data
        special_cases_data = []
        
        if not threshold_special_cases.empty:
            threshold_cases_processed = threshold_special_cases.groupby('HWB').agg({
                'CONSIGNEE_NAME': 'first',
                'CONSIGNEE_ZIP': 'first',
                'MATCHED_ROUTE': 'first',
                'WEIGHT': 'max',
                'VOLUMETRIC_WEIGHT': 'max',
                'PIECES': 'first'
            }).reset_index()
            
            def get_trigger_reason(row):
                reasons = []
                if row['WEIGHT'] > weight_thr: reasons.append(f'Weight >{weight_thr}kg')
                if row['VOLUMETRIC_WEIGHT'] > vol_weight_thr: reasons.append(f'Volumetric >{vol_weight_thr}kg')
                if row['PIECES'] > pieces_thr: reasons.append(f'Pieces >{pieces_thr}')
                return ', '.join(reasons) if reasons else None
            
            threshold_cases_processed['TRIGGER_REASON'] = threshold_cases_processed.apply(get_trigger_reason, axis=1)
            threshold_cases_processed['WEIGHT_PER_PIECE'] = threshold_cases_processed.apply(
                lambda x: round(x['WEIGHT']/x['PIECES'], 2) if x['PIECES'] > pieces_thr else None, axis=1)
            
            def get_vehicle_suggestion(row):
                conditions = [
                    row['WEIGHT'] > vehicle_weight_thr,
                    row['VOLUMETRIC_WEIGHT'] > vehicle_vol_thr,
                    row['PIECES'] > vehicle_pieces_thr,
                    (not pd.isna(row['WEIGHT_PER_PIECE'])) and 
                    (row['WEIGHT_PER_PIECE'] > vehicle_kg_per_piece_thr) and 
                    (row['PIECES'] > vehicle_van_max_pieces)
                ]
                return "Truck" if any(conditions) else "Van"
            
            threshold_cases_processed['Capacity_Suggestion'] = threshold_cases_processed.apply(get_vehicle_suggestion, axis=1)
            threshold_cases_processed = threshold_cases_processed.sort_values(by="CONSIGNEE_ZIP", ascending=True)
            special_cases_data.append(('Threshold Cases', threshold_cases_processed))
        
        # Add multi-shipment data
        if not multi_shipment_special.empty:
            multi_shipment_special = multi_shipment_special.sort_values(by=['zip_code', 'total_shipments'], ascending=[True, False])
            multi_shipment_special['TRIGGER_REASON'] = multi_shipment_special.apply(
                lambda x: f'Multiple Shipments ({x["total_shipments"]})', axis=1)
            # Rename columns to match threshold cases
            multi_shipment_special = multi_shipment_special.rename(columns={
                'consignee_name': 'CONSIGNEE_NAME',
                'zip_code': 'CONSIGNEE_ZIP',
                'matched_route': 'MATCHED_ROUTE',
                'total_pieces': 'PIECES'
            })
            multi_shipment_special['HWB'] = ''
            multi_shipment_special['WEIGHT'] = 0
            multi_shipment_special['VOLUMETRIC_WEIGHT'] = 0
            multi_shipment_special['WEIGHT_PER_PIECE'] = None
            multi_shipment_special['Capacity_Suggestion'] = 'Multiple Shipments'
            special_cases_data.append(('Multi-Shipment Cases', multi_shipment_special))
        
        # Create Excel file using pd.ExcelWriter
        with pd.ExcelWriter(special_cases_path, engine='openpyxl') as writer:
            if special_cases_data:
                # Write first dataset (threshold cases)
                first_data = special_cases_data[0][1]
                first_data.to_excel(writer, index=False, startrow=5, sheet_name='Special Cases')
                workbook = writer.book
                sheet = writer.sheets['Special Cases']
                add_dhl_branding_to_excel(workbook, sheet, "DHL Special Cases Report")
                
                # Add multi-shipment data below if exists
                if len(special_cases_data) > 1:
                    current_row = sheet.max_row + 3
                    sheet.cell(row=current_row, column=1, value=f"Multiple Shipments Customers (≥{multi_shipment_thr} shipments)")
                    sheet.cell(row=current_row, column=1).font = Font(bold=True)
                    current_row += 2
                    
                    multi_data = special_cases_data[1][1]
                    for idx, row in multi_data.iterrows():
                        sheet.cell(row=current_row, column=1, value=row['HWB'])
                        sheet.cell(row=current_row, column=2, value=row['CONSIGNEE_NAME'])
                        sheet.cell(row=current_row, column=3, value=row['CONSIGNEE_ZIP'])
                        sheet.cell(row=current_row, column=4, value=row['MATCHED_ROUTE'])
                        sheet.cell(row=current_row, column=7, value=row['PIECES'])
                        sheet.cell(row=current_row, column=10, value=row['TRIGGER_REASON'])
                        current_row += 1
                
                auto_adjust_column_width(sheet)
            else:
                # Empty file
                pd.DataFrame(columns=['HWB', 'CONSIGNEE_NAME', 'TRIGGER_REASON']).to_excel(writer, index=False, startrow=5)
                workbook = writer.book
                sheet = writer.sheets['Sheet1']
                add_dhl_branding_to_excel(workbook, sheet, "DHL Special Cases Report")
                sheet.cell(row=6, column=1, value="No special cases found")
        
        st.success("✅ Special cases report created")
    except Exception as e:
        st.error(f"❌ Failed to create special cases report: {str(e)}")

    # 4. MULTIPLE SHIPMENTS REPORT
    try:
        st.write("📊 Creating multiple shipments report...")
        multi_shipment_customers = identify_multi_shipment_customers(manifest_df)
        if not multi_shipment_customers.empty:
            multi_shipment_customers.rename(columns={
                'total_shipments': 'Shipment Count',
                'total_pieces': 'Total Pieces',
                'zip_code': 'ZIP'
            }, inplace=True)
            
            multi_shipment_customers = multi_shipment_customers[['CONSIGNEE_NAME_NORM', 'ZIP', 'Shipment Count', 'Total Pieces']]
            
            with pd.ExcelWriter(multi_shipments_path, engine='openpyxl') as writer:
                multi_shipment_customers.to_excel(writer, index=False, startrow=5)
                workbook = writer.book
                sheet = writer.sheets['Sheet1']
                add_dhl_branding_to_excel(workbook, sheet, "DHL Multiple Shipments Report")
                auto_adjust_column_width(sheet)
        else:
            with pd.ExcelWriter(multi_shipments_path, engine='openpyxl') as writer:
                pd.DataFrame(columns=['CONSIGNEE_NAME_NORM', 'ZIP', 'Shipment Count', 'Total Pieces']).to_excel(writer, index=False, startrow=5)
                workbook = writer.book
                sheet = writer.sheets['Sheet1']
                add_dhl_branding_to_excel(workbook, sheet, "DHL Multiple Shipments Report")
        st.success("✅ Multiple shipments report created")
    except Exception as e:
        st.error(f"❌ Failed to create multiple shipments report: {str(e)}")

    # 5. MATCHING DETAILS
    try:
        st.write("🔍 Creating matching details report...")
        matching_details = manifest_df[['HWB', 'CONSIGNEE_NAME', 'CONSIGNEE_ZIP', 'CONSIGNEE_ADDRESS', 'MATCHED_ROUTE', 'MATCH_METHOD']].copy()
        matching_details['MATCHED_ROUTE'] = matching_details['MATCHED_ROUTE'].fillna('UNMATCHED')
        matching_details['MATCH_METHOD'] = matching_details['MATCH_METHOD'].fillna('UNMATCHED')
        matching_details.rename(columns={'MATCH_METHOD': 'MATCHING_METHOD'}, inplace=True)
        
        with pd.ExcelWriter(matching_details_path, engine='openpyxl') as writer:
            matching_details.to_excel(writer, index=False, startrow=5)
            workbook = writer.book
            sheet = writer.sheets['Sheet1']
            add_dhl_branding_to_excel(workbook, sheet, "DHL Route Matching Details")
            auto_adjust_column_width(sheet)
        st.success("✅ Matching details report created")
    except Exception as e:
        st.error(f"❌ Failed to create matching details report: {str(e)}")

    # 6. WTH MPCS REPORT
    try:
        st.write("📦 Creating WTH MPCS report...")
        
        # Prepare WTH MPCS data
        wth_data = []
        
        if not threshold_special_cases.empty:
            threshold_cases_for_wth = threshold_special_cases.groupby('HWB').agg({
                'CONSIGNEE_NAME': 'first',
                'CONSIGNEE_ZIP': 'first',
                'CONSIGNEE_CITY': 'first',
                'WEIGHT': 'max',
                'VOLUMETRIC_WEIGHT': 'max',
                'PIECES': 'first'
            }).reset_index()
            
            def get_trigger_reason_wth(row):
                reasons = []
                if row['WEIGHT'] > weight_thr: reasons.append(f'Weight >{weight_thr}kg')
                if row['VOLUMETRIC_WEIGHT'] > vol_weight_thr: reasons.append(f'Volumetric >{vol_weight_thr}kg')
                if row['PIECES'] > pieces_thr: reasons.append(f'Pieces >{pieces_thr}')
                return ', '.join(reasons) if reasons else None
            
            threshold_cases_for_wth['TRIGGER_REASON'] = threshold_cases_for_wth.apply(get_trigger_reason_wth, axis=1)
            threshold_cases_for_wth = threshold_cases_for_wth.sort_values(by="CONSIGNEE_ZIP", ascending=True)
            wth_data.append(threshold_cases_for_wth)
        
        if not multi_shipment_special.empty:
            multi_with_city = manifest_df.groupby('CONSIGNEE_NAME_NORM').agg(
                total_shipments=('HWB', 'nunique'),
                total_pieces=('PIECES', 'sum'),
                total_weight=('WEIGHT', 'sum'),
                total_vol_weight=('VOLUMETRIC_WEIGHT', 'sum'),
                zip_code=('CONSIGNEE_ZIP', 'first'),
                consignee_name=('CONSIGNEE_NAME', 'first'),
                city=('CONSIGNEE_CITY', 'first')
            ).reset_index()
            
            multi_for_wth = multi_with_city[multi_with_city['total_shipments'] >= multi_shipment_thr].copy()
            multi_for_wth = multi_for_wth.sort_values(by=['zip_code', 'total_shipments'], ascending=[True, False])
            
            # Rename columns to match threshold format
            multi_for_wth = multi_for_wth.rename(columns={
                'consignee_name': 'CONSIGNEE_NAME',
                'zip_code': 'CONSIGNEE_ZIP',
                'city': 'CONSIGNEE_CITY',
                'total_pieces': 'PIECES',
                'total_weight': 'WEIGHT',
                'total_vol_weight': 'VOLUMETRIC_WEIGHT'
            })
            multi_for_wth['HWB'] = ''
            multi_for_wth['TRIGGER_REASON'] = multi_for_wth.apply(
                lambda x: f'Multiple Shipments ≥{multi_shipment_thr} ({x["total_shipments"]})', axis=1)
            
            wth_data.append(multi_for_wth)
        
        # Create WTH MPCS Excel file
        with pd.ExcelWriter(wth_mpcs_path, engine='openpyxl') as writer:
            if wth_data:
                # Combine all WTH data
                combined_wth = pd.concat(wth_data, ignore_index=True)
                wth_columns = ['CONSIGNEE_NAME', 'CONSIGNEE_ZIP', 'CONSIGNEE_CITY', 'PIECES', 'WEIGHT', 'VOLUMETRIC_WEIGHT', 'HWB', 'TRIGGER_REASON']
                combined_wth = combined_wth[wth_columns]
                combined_wth.to_excel(writer, index=False, startrow=5)
                workbook = writer.book
                sheet = writer.sheets['Sheet1']
                add_dhl_branding_to_excel(workbook, sheet, "DHL WTH MPCS Report")
                auto_adjust_column_width(sheet)
            else:
                pd.DataFrame(columns=['CONSIGNEE_NAME', 'ZIP', 'CITY', 'PIECES', 'WEIGHT', 'VOLUMETRIC_WEIGHT', 'HWB', 'TRIGGER_REASON']).to_excel(writer, index=False, startrow=5)
                workbook = writer.book
                sheet = writer.sheets['Sheet1']
                add_dhl_branding_to_excel(workbook, sheet, "DHL WTH MPCS Report")
                sheet.cell(row=6, column=1, value="No special cases found")
        
        st.success("✅ WTH MPCS report created")
    except Exception as e:
        st.error(f"❌ Failed to create WTH MPCS report: {str(e)}")

    # 7. PRIORITY SHIPMENTS - ENHANCED WITH TDL (10:30) AND TDX (12:00)
    try:
        st.write("🚨 Creating priority shipments report...")
        if 'PCC' in manifest_df.columns:
            manifest_df['PCC'] = manifest_df['PCC'].astype(str).str.strip().str.upper()
            # Updated priority codes to include TDL in correct order: CMX, TDL, TDX
            priority_codes = ['CMX', 'TDL', 'TDX']
            priority_pccs = manifest_df[manifest_df['PCC'].isin(priority_codes)]
            
            if not priority_pccs.empty:
                # Group by priority level - CMX stays at top, TDL in middle, TDX at bottom
                group1 = priority_pccs[priority_pccs['PCC'] == 'CMX'].sort_values(
                    by=['CONSIGNEE_ZIP', 'MATCHED_ROUTE'], ascending=[True, True])
                group2 = priority_pccs[priority_pccs['PCC'] == 'TDL'].sort_values(
                    by=['CONSIGNEE_ZIP', 'MATCHED_ROUTE'], ascending=[True, True])
                group3 = priority_pccs[priority_pccs['PCC'] == 'TDX'].sort_values(
                    by=['CONSIGNEE_ZIP', 'MATCHED_ROUTE'], ascending=[True, True])
                
                cols = ['MATCHED_ROUTE', 'HWB', 'CONSIGNEE_NAME', 'CONSIGNEE_ZIP', 'PCC', 'WEIGHT', 'VOLUMETRIC_WEIGHT', 'PIECES']
                
                with pd.ExcelWriter(priority_path, engine='openpyxl') as writer:
                    current_row = 6
                    workbook = None
                    sheet = None
                    
                    # CMX Priority Shipments (Highest Priority)
                    if not group1.empty:
                        group1[cols].to_excel(writer, index=False, startrow=current_row+2, sheet_name='Priority Shipments')
                        workbook = writer.book
                        sheet = writer.sheets['Priority Shipments']
                        add_dhl_branding_to_excel(workbook, sheet, "DHL Priority Shipments Report")
                        
                        sheet.cell(row=current_row, column=1, value="CMX Priority Shipments")
                        sheet.cell(row=current_row, column=1).font = Font(bold=True)
                        current_row = sheet.max_row + 2
                    
                    # TDL Priority Shipments (10:30) - NEW PRODUCT
                    if not group2.empty:
                        if workbook is None:
                            group2[cols].to_excel(writer, index=False, startrow=current_row+2, sheet_name='Priority Shipments')
                            workbook = writer.book
                            sheet = writer.sheets['Priority Shipments']
                            add_dhl_branding_to_excel(workbook, sheet, "DHL Priority Shipments Report")
                        
                        sheet.cell(row=current_row, column=1, value="TDL Priority Shipments (10:30)")
                        sheet.cell(row=current_row, column=1).font = Font(bold=True)
                        current_row += 2
                        
                        # Write headers
                        for col_idx, col in enumerate(cols, 1):
                            sheet.cell(row=current_row, column=col_idx, value=col)
                            sheet.cell(row=current_row, column=col_idx).font = Font(bold=True)
                        current_row += 1
                        
                        # Write data
                        for _, row in group2[cols].iterrows():
                            for col_idx, value in enumerate(row, 1):
                                sheet.cell(row=current_row, column=col_idx, value=value)
                            current_row += 1
                        current_row += 2
                    
                    # TDX Priority Shipments (12:00) - UPDATED WITH TIME
                    if not group3.empty:
                        if workbook is None:
                            group3[cols].to_excel(writer, index=False, startrow=current_row+2, sheet_name='Priority Shipments')
                            workbook = writer.book
                            sheet = writer.sheets['Priority Shipments']
                            add_dhl_branding_to_excel(workbook, sheet, "DHL Priority Shipments Report")
                        
                        sheet.cell(row=current_row, column=1, value="TDX Priority Shipments (12:00)")
                        sheet.cell(row=current_row, column=1).font = Font(bold=True)
                        current_row += 2
                        
                        # Write headers
                        for col_idx, col in enumerate(cols, 1):
                            sheet.cell(row=current_row, column=col_idx, value=col)
                            sheet.cell(row=current_row, column=col_idx).font = Font(bold=True)
                        current_row += 1
                        
                        # Write data
                        for _, row in group3[cols].iterrows():
                            for col_idx, value in enumerate(row, 1):
                                sheet.cell(row=current_row, column=col_idx, value=value)
                            current_row += 1
                    
                    if workbook and sheet:
                        auto_adjust_column_width(sheet)
                    else:
                        # No priority shipments found
                        pd.DataFrame(columns=cols).to_excel(writer, index=False, startrow=5)
                        workbook = writer.book
                        sheet = writer.sheets['Sheet1']
                        add_dhl_branding_to_excel(workbook, sheet, "DHL Priority Shipments Report")
                        sheet.cell(row=6, column=1, value="No priority shipments found")
            else:
                with pd.ExcelWriter(priority_path, engine='openpyxl') as writer:
                    pd.DataFrame(columns=['MATCHED_ROUTE', 'HWB', 'CONSIGNEE_NAME', 'PCC']).to_excel(writer, index=False, startrow=5)
                    workbook = writer.book
                    sheet = writer.sheets['Sheet1']
                    add_dhl_branding_to_excel(workbook, sheet, "DHL Priority Shipments Report")
                    sheet.cell(row=6, column=1, value="No priority shipments found")
        else:
            with pd.ExcelWriter(priority_path, engine='openpyxl') as writer:
                pd.DataFrame(columns=['MATCHED_ROUTE', 'HWB', 'CONSIGNEE_NAME', 'PCC']).to_excel(writer, index=False, startrow=5)
                workbook = writer.book
                sheet = writer.sheets['Sheet1']
                add_dhl_branding_to_excel(workbook, sheet, "DHL Priority Shipments Report")
                sheet.cell(row=6, column=1, value="No PCC data available")
        
        st.success("✅ Priority shipments report created")
    except Exception as e:
        st.error(f"❌ Failed to create priority shipments report: {str(e)}")

    # RETURN ALL RESULTS - THIS IS THE ONLY RETURN STATEMENT IN THE FUNCTION
    return timestamp, route_summary, specialized_reports, multi_shipments_path, targets_df
        # ───────────────────────────
        #  CONT. MULTI-SHIPMENTS FILE
        # ───────────────────────────
        multi_shipment_customers = identify_multi_shipment_customers(manifest_df)

        if not multi_shipment_customers.empty:
            multi_shipment_customers.rename(
                columns={
                    'total_shipments': 'Shipment Count',
                    'total_pieces':   'Total Pieces',
                    'zip_code':       'ZIP'
                },
                inplace=True
            )

            cols_order = ['CONSIGNEE_NAME_NORM', 'ZIP', 'Shipment Count', 'Total Pieces']
            multi_shipment_customers = multi_shipment_customers[cols_order]

            with pd.ExcelWriter(multi_shipments_path, engine='openpyxl') as writer:
                multi_shipment_customers.to_excel(writer, index=False, startrow=5)
                wb_ms  = writer.book
                ws_ms  = writer.sheets['Sheet1']
                add_dhl_branding_to_excel(wb_ms, ws_ms, "DHL Multiple Shipments Report")
                auto_adjust_column_width(ws_ms)
        else:
            with pd.ExcelWriter(multi_shipments_path, engine='openpyxl') as writer:
                pd.DataFrame(
                    columns=['CONSIGNEE_NAME_NORM', 'ZIP', 'Shipment Count', 'Total Pieces']
                ).to_excel(writer, index=False, startrow=5)
                wb_ms  = writer.book
                ws_ms  = writer.sheets['Sheet1']
                add_dhl_branding_to_excel(wb_ms, ws_ms, "DHL Multiple Shipments Report")

        # ───────────────────────────
        #        MATCHING DETAILS
        # ───────────────────────────
        matching_details = manifest_df[
            ['HWB', 'CONSIGNEE_NAME', 'CONSIGNEE_ZIP',
             'CONSIGNEE_ADDRESS', 'MATCHED_ROUTE', 'MATCH_METHOD']
        ].copy()

        matching_details.fillna({'MATCHED_ROUTE': 'UNMATCHED',
                                 'MATCH_METHOD': 'UNMATCHED'}, inplace=True)
        matching_details.rename(columns={'MATCH_METHOD': 'MATCHING_METHOD'}, inplace=True)

        with pd.ExcelWriter(matching_details_path, engine='openpyxl') as writer:
            matching_details.to_excel(writer, index=False, startrow=5)
            wb_md = writer.book
            ws_md = writer.sheets['Sheet1']
            add_dhl_branding_to_excel(wb_md, ws_md, "DHL Route Matching Details")
            auto_adjust_column_width(ws_md)

        # ───────────────────────────
        #           WTH / MPCS
        # ───────────────────────────
        wb_wth = Workbook()
        ws_wth = wb_wth.active
        ws_wth.title = "WTH MPCS Report"
        add_dhl_branding_to_excel(wb_wth, ws_wth, "DHL WTH MPCS Report")

        header_row = 6
        headers = ['CONSIGNEE_NAME', 'ZIP', 'CITY', 'PIECES', 'WEIGHT',
                   'VOLUMETRIC_WEIGHT', 'HWB', '', '', 'TRIGGER_REASON']
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws_wth.cell(row=header_row, column=col_idx, value=hdr)
            cell.font = Font(bold=True)

        wth_row = header_row + 1

        # ► threshold special cases
        if not threshold_special_cases.empty:
            group = threshold_special_cases.groupby('HWB').agg({
                'CONSIGNEE_NAME':    'first',
                'CONSIGNEE_ZIP':     'first',
                'CONSIGNEE_CITY':    'first',
                'WEIGHT':            'max',
                'VOLUMETRIC_WEIGHT': 'max',
                'PIECES':            'first'
            }).reset_index()

            def trigger(r):
                reason = []
                if r['WEIGHT'] > weight_thr:               reason.append(f'Weight>{weight_thr}')
                if r['VOLUMETRIC_WEIGHT'] > vol_weight_thr:reason.append(f'Vol>{vol_weight_thr}')
                if r['PIECES'] > pieces_thr:               reason.append(f'Pieces>{pieces_thr}')
                return ', '.join(reason)

            group['TRIGGER_REASON'] = group.apply(trigger, axis=1)

            for _, r in group.sort_values('CONSIGNEE_ZIP').iterrows():
                ws_wth.cell(wth_row, 1, r['CONSIGNEE_NAME'])
                ws_wth.cell(wth_row, 2, r['CONSIGNEE_ZIP'])
                ws_wth.cell(wth_row, 3, r['CONSIGNEE_CITY'])
                ws_wth.cell(wth_row, 4, r['PIECES'])
                ws_wth.cell(wth_row, 5, r['WEIGHT'])
                ws_wth.cell(wth_row, 6, r['VOLUMETRIC_WEIGHT'])
                ws_wth.cell(wth_row, 7, r['HWB'])
                ws_wth.cell(wth_row,10, r['TRIGGER_REASON'])
                wth_row += 1

        # ► multi-shipment customers
        if not multi_shipment_special.empty:
            multi_city = manifest_df.groupby('CONSIGNEE_NAME_NORM').agg(
                total_shipments=('HWB', 'nunique'),
                total_pieces   =('PIECES', 'sum'),
                total_weight   =('WEIGHT', 'sum'),
                total_vol      =('VOLUMETRIC_WEIGHT', 'sum'),
                zip_code       =('CONSIGNEE_ZIP', 'first'),
                name           =('CONSIGNEE_NAME', 'first'),
                city           =('CONSIGNEE_CITY', 'first')
            ).reset_index()

            multi_city = multi_city[multi_city['total_shipments'] >= multi_shipment_thr]

            for _, r in multi_city.sort_values(['zip_code', 'total_shipments'], ascending=[True, False]).iterrows():
                ws_wth.cell(wth_row, 1, r['name'])
                ws_wth.cell(wth_row, 2, r['zip_code'])
                ws_wth.cell(wth_row, 3, r['city'])
                ws_wth.cell(wth_row, 4, r['total_pieces'])
                ws_wth.cell(wth_row, 5, round(r['total_weight'], 1))
                ws_wth.cell(wth_row, 6, round(r['total_vol'], 1))
                ws_wth.cell(wth_row,10, f'Multiple Shipments ≥{multi_shipment_thr} ({r["total_shipments"]})')
                wth_row += 1

        if wth_row == header_row + 1:
            ws_wth.cell(header_row + 1, 1, "No special cases found")

        auto_adjust_column_width(ws_wth)
        wb_wth.save(wth_mpcs_path)

        # ───────────────────────────
        #       PRIORITY SHIPMENTS
        # ───────────────────────────
        if 'PCC' in manifest_df.columns:
            manifest_df['PCC'] = manifest_df['PCC'].astype(str).str.upper()
            pr_codes = ['CMX', 'WMX', 'TDT', 'TDY']
            pr_df = manifest_df[manifest_df['PCC'].isin(pr_codes)]

            wb_pr = Workbook()
            ws_pr = wb_pr.active
            ws_pr.title = "Priority Shipments"
            add_dhl_branding_to_excel(wb_pr, ws_pr, "DHL Priority Shipments Report")

            pr_sections = [('CMX/WMX Priority Shipments', ['CMX', 'WMX']),
                           ('TDT/TDY Priority Shipments', ['TDT', 'TDY'])]

            pr_row = 6
            cols = ['MATCHED_ROUTE', 'HWB', 'CONSIGNEE_NAME', 'CONSIGNEE_ZIP',
                    'PCC', 'WEIGHT', 'VOLUMETRIC_WEIGHT', 'PIECES']

            for title, codes in pr_sections:
                subset = pr_df[pr_df['PCC'].isin(codes)]
                if subset.empty:
                    continue

                ws_pr.cell(pr_row, 1, title).font = Font(bold=True)
                pr_row += 2

                for c_idx, c in enumerate(cols, 1):
                    ws_pr.cell(pr_row, c_idx, c).font = Font(bold=True)
                pr_row += 1

                for r_idx, row in enumerate(
                        dataframe_to_rows(subset[cols], index=False, header=False), pr_row):
                    for c_idx, val in enumerate(row, 1):
                        ws_pr.cell(r_idx, c_idx, val)
                pr_row = ws_pr.max_row + 3

            auto_adjust_column_width(ws_pr)
            wb_pr.save(priority_path)
        else:
            wb_pr = Workbook()
            ws_pr = wb_pr.active
            add_dhl_branding_to_excel(wb_pr, ws_pr, "DHL Priority Shipments Report")
            ws_pr.cell(6, 1, "No PCC data available")
            wb_pr.save(priority_path)

        # ───────────────────────────
        #  RETURN FOR STREAMLIT MAIN
        # ───────────────────────────
        return timestamp, route_summary, specialized_reports, multi_shipments_path, targets_df
